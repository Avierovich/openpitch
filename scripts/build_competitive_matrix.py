"""Generate docs/competitive-matrix.xlsx.

This workbook is a PM/PMM artifact, not runtime product code. It deliberately
mixes direct incumbents, open/MCP adjacencies, and regional/private-market
platforms so OpenPitch's wedge is judged against the right competitive set.

Reproducible:
    python scripts/build_competitive_matrix.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "docs" / "competitive-matrix.xlsx"


# ── Competitor universe ──────────────────────────────────────────────────────

CORE_COMPETITORS = [
    "PitchBook",
    "CB Insights",
    "Crunchbase",
    "Dealroom",
    "Tracxn",
    "Harmonic",
    "MAGNiTT",
    "Wamda",
    "Sacra",
    "OpenBB",
    "AI Funding Tracker",
    "Crunchbase MCP",
    "OpenPitch",
]


# ── Styling ──────────────────────────────────────────────────────────────────

NAVY = "1F2A44"
ACCENT = "2E7D6F"
LIGHT_ACCENT = "E3F2EF"
LIGHT_BLUE = "EAF2FF"
LIGHT_GRAY = "F4F6F8"
YES_FILL = PatternFill("solid", fgColor="C8E6C9")
PARTIAL_FILL = PatternFill("solid", fgColor="FFF3C4")
NO_FILL = PatternFill("solid", fgColor="F5D5D0")
US_FILL = PatternFill("solid", fgColor=LIGHT_ACCENT)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=NAVY, size=14)
SUBTITLE_FONT = Font(italic=True, color="666666", size=10)
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=ACCENT)
NOTE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_header(ws, ncols: int, row: int = 1, *, fill=HDR_FILL) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER


def _grid(ws, nrows: int, ncols: int) -> None:
    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT


def _mark_fill(ws, header_row: list[str], start_row: int = 2) -> None:
    """Color ✓/◐/✗ cells and shade the OpenPitch column."""
    us_col = header_row.index("OpenPitch") + 1 if "OpenPitch" in header_row else None
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            v = str(cell.value).strip() if cell.value is not None else ""
            if v == "✓":
                cell.fill = YES_FILL
                cell.alignment = CENTER
            elif v == "◐":
                cell.fill = PARTIAL_FILL
                cell.alignment = CENTER
            elif v == "✗":
                cell.fill = NO_FILL
                cell.alignment = CENTER
            elif us_col and cell.column == us_col:
                cell.fill = US_FILL
            elif cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FBFBFB")
            if len(v) <= 3:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT


def _set_widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_sheet(wb, title: str, rows: list[list], widths: list[int], *, freeze: str = "B2") -> None:
    ws = wb.create_sheet(title)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    _grid(ws, len(rows), max(len(r) for r in rows))
    _style_header(ws, len(rows[0]))
    _set_widths(ws, widths)
    ws.freeze_panes = freeze
    if title in {"Core Overview", "Feature Matrix", "Agent/Open Landscape", "OSS + Free Direct Landscape"}:
        _mark_fill(ws, rows[0])
    return ws


# ── Sheet 1: executive summary ───────────────────────────────────────────────

EXECUTIVE_SUMMARY = [
    ["OpenPitch Competitive Matrix — Verified PM/PMM View", "", "", ""],
    ["Updated", "2026-06-13", "Workbook owner", "OpenPitch"],
    ["Bottom line", "The market is more crowded and more agent-native than the original matrix implied. OpenPitch's defensible wedge is not 'startup data in an agent'; it is free/open, no-key, sourced/confidence-scored AI-startup intelligence with contradiction handling.", "", ""],
    ["Strongest direct threats", "CB Insights (agent workforce), Dealroom (startup ecosystem data + MCP), Exa Company Researcher (open company-research agent), Octagon Funding MCP, Sieve MCP, CompanyLens MCP, Crunchbase wrappers, OpenBB (open financial-data MCP distribution)", "", ""],
    ["OpenPitch must win on", "1) zero-config local install; 2) per-figure provenance/confidence; 3) contradiction detection; 4) AI-startup focus; 5) public corrections/auditability", "", ""],
    ["Launch implication", "Do not launch as a broad PitchBook clone. Launch with a dense 10-company seed, strong public-source discrepancies, and a working MCP demo.", "", ""],
    ["Legend", "✓ = yes / strong; ◐ = partial / limited / indirect; ✗ = no / not visible publicly", "", ""],
]


# ── Sheet 2: core overview ───────────────────────────────────────────────────

CORE_OVERVIEW = [
    ["Dimension", *CORE_COMPETITORS],
    [
        "Type",
        "Institutional private-market data",
        "Predictive market intelligence",
        "Broad company/funding DB",
        "Startup ecosystem data platform",
        "Private-market intelligence",
        "Startup sourcing graph",
        "Emerging-markets private-capital data",
        "MENA startup media/research",
        "Private-company research + revenue data",
        "Open financial data/workspace",
        "AI funding media/tracker",
        "Open-source MCP wrapper",
        "Open signal layer",
    ],
    [
        "Primary region/scope",
        "Global",
        "Global",
        "Global",
        "Global tech ecosystems",
        "Global, strong EM/India footprint",
        "Global startup/company graph",
        "MENA, Africa, Turkey, Pakistan, SEA",
        "MENA",
        "Selective private companies",
        "Public markets / finance workflows",
        "AI startups",
        "Crunchbase API coverage",
        "Global AI + MENA-AI seed",
    ],
    [
        "Pricing/access",
        "Quote-only; high annual contract",
        "Quote-only enterprise",
        "Freemium + paid plans/API",
        "Demo/API sales motion",
        "Lite/free limited + premium/API add-ons",
        "Quote-only",
        "Quote/demo; reports",
        "Free media + paid reports",
        "Subscription tiers; API/MCP on higher tiers reported",
        "Open-source plus Workspace/account/provider keys",
        "Free website",
        "Open-source code; requires Crunchbase access/API",
        "$0 open source",
    ],
    [
        "Free to use without paid data key",
        "✗",
        "✗",
        "◐",
        "✗",
        "◐",
        "✗",
        "✗",
        "✓",
        "◐",
        "◐",
        "✓",
        "✗",
        "✓",
    ],
    [
        "Agent/MCP posture",
        "◐",
        "✓",
        "◐",
        "✓",
        "◐",
        "◐",
        "✗",
        "✗",
        "◐",
        "✓",
        "✗",
        "✓",
        "✓",
    ],
    [
        "ARR/revenue estimates",
        "◐",
        "◐",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✓",
        "✗",
        "◐",
        "✗",
        "✓",
    ],
    [
        "Per-figure provenance + confidence",
        "◐",
        "◐",
        "◐",
        "◐",
        "◐",
        "✗",
        "◐",
        "◐",
        "◐",
        "◐",
        "◐",
        "✗",
        "✓",
    ],
    [
        "Contradiction detection",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✓",
    ],
    [
        "Open source / open data posture",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✗",
        "✓",
        "✗",
        "✓ code only",
        "✓",
    ],
    [
        "OpenPitch threat level",
        "Low-medium",
        "High",
        "Medium",
        "High",
        "Medium",
        "Medium",
        "Medium in MENA",
        "Low-medium",
        "Medium for revenue data",
        "Medium as open/MCP incumbent",
        "Low-medium content threat",
        "Low wrapper threat",
        "Self",
    ],
]


# ── Sheet 3: feature matrix ──────────────────────────────────────────────────

FEATURE_MATRIX = [
    ["Feature", *CORE_COMPETITORS],
    ["Funding-round data", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "◐", "◐", "✗", "✓", "✓", "✓"],
    ["Valuations", "✓", "✓", "◐", "✓", "✓", "◐", "◐", "✗", "◐", "✗", "◐", "◐", "✓"],
    ["ARR / revenue estimates", "◐", "◐", "✗", "✗", "✗", "✗", "✗", "✗", "✓", "✗", "◐", "✗", "✓"],
    ["Verified financial statements", "✓", "◐", "✗", "✗", "✗", "✗", "◐", "✗", "◐", "✓ public only", "✗", "✗", "✗"],
    ["Headcount / hiring signals", "✓", "✓", "◐", "✓", "✓", "✓", "◐", "✗", "✗", "✗", "✗", "◐", "✓"],
    ["People / founder graph", "◐", "◐", "◐", "◐", "◐", "✓", "◐", "✗", "✗", "✗", "✗", "◐", "✗"],
    ["Podcast / alt-data mining", "✗", "◐", "✗", "✗", "✗", "◐", "✗", "✗", "✗", "✗", "✗", "✗", "✓"],
    ["Source citations per figure", "◐", "◐", "◐", "◐", "◐", "✗", "◐", "◐", "◐", "◐", "◐", "✗", "✓"],
    ["Confidence scoring visible", "✗", "◐", "✗", "◐", "◐", "◐", "✗", "✗", "◐", "✗", "✗", "✗", "✓"],
    ["Contradiction detection", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✓"],
    ["Daily/frequent freshness", "◐", "◐", "◐", "✓", "✓", "✓", "◐", "◐", "◐", "◐", "✓", "◐", "✓"],
    ["MCP / agent-native access", "◐", "✓", "◐", "✓", "◐", "◐", "✗", "✗", "◐", "✓", "✗", "✓", "✓"],
    ["Public API / data feed", "✓", "✓", "✓", "✓", "✓", "✓", "◐", "✗", "✓", "✓", "✗", "✓ via CB", "✓"],
    ["Event / webhook feed", "◐", "◐", "◐", "◐", "◐", "✓", "✗", "✗", "✗", "◐", "✗", "✗", "✓"],
    ["MENA coverage", "◐", "◐", "◐", "◐", "◐", "◐", "✓", "✓", "✗", "✗", "✗", "◐", "✓"],
    ["AI-startup focus", "◐", "◐", "◐", "◐", "◐", "◐", "◐", "◐", "◐", "✗", "✓", "◐", "✓"],
    ["Free/open source", "✗", "✗", "◐", "✗", "◐", "✗", "✗", "◐", "✗", "✓", "✗", "✓ code", "✓"],
]


# ── Sheet 4: pricing and access ──────────────────────────────────────────────

PRICING = [
    ["Provider", "Public pricing posture", "Free tier", "Agent/API access", "What pricing implies for OpenPitch"],
    ["PitchBook", "Not public; high annual contracts reported by third parties", "No", "API/data feed; AI capabilities; ChatGPT/Premium Connector public references", "OpenPitch should not compete on diligence-grade breadth; win no-key first-look."],
    ["CB Insights", "Quote-only enterprise", "No", "Team of Agents; Microsoft 365 Copilot connector; MCP-compatible posture publicly claimed", "This is the strongest signal that incumbents are agent-native; OpenPitch must emphasize free/open/transparency."],
    ["Crunchbase", "Freemium plus paid/API plans", "Limited", "API; AI predictions/insights; third-party MCP wrappers", "Sets the 'good enough for free' baseline but lacks ARR/confidence/contradictions."],
    ["Dealroom", "Demo/sales motion; API", "No broad free tier", "API + MCP support publicly documented in release notes", "Very important competitor for ecosystem/startup data; OpenPitch must differentiate on no-key and confidence."],
    ["Tracxn", "Lite/free limited; premium/API add-ons", "Limited", "API/datadump add-ons", "Broad private-market data; less agent-native in public messaging."],
    ["Harmonic", "Quote-only", "Limited/no broad free", "API/AI search/alerts; agent-ish", "Threat for sourcing/team data, not ARR/provenance."],
    ["MAGNiTT", "Pricing page with sales motion", "Limited reports/content", "Platform/research; API not prominent publicly", "Strong MENA private-capital source; OpenPitch should be honest that MENA confidence starts lower."],
    ["Wamda", "Free media/research; paid reports possible", "Yes", "No visible API/MCP", "Content source more than platform competitor."],
    ["Sacra", "Subscription; third-party listings mention paid MCP/API tiers", "Limited content", "Company API; MCP access reported on pricing listings", "Closest to private-company revenue intelligence; narrower coverage and not open."],
    ["OpenBB", "Open-source plus Workspace/account/provider model", "Yes, with caveats", "MCP server and Workspace MCP tools", "Most important open/MCP adjacent; could become partner/provider path."],
    ["AI Funding Tracker", "Free website", "Yes", "No API/MCP visible", "Content/SEO competitor; machine-consumption gap is OpenPitch's advantage."],
    ["Crunchbase MCP", "Open-source wrapper", "No if CB data requires paid access", "MCP server code", "Validates category but is a thin wrapper over proprietary data."],
    ["OpenPitch", "$0 open source", "Yes", "Local MCP, raw JSON, events planned", "Must preserve no-key install and transparent audit trail."],
]


# ── Sheet 5: agent/open landscape ────────────────────────────────────────────

AGENT_OPEN = [
    ["Capability", "OpenBB", "Dealroom", "CB Insights", "PitchBook", "Crunchbase MCP", "Sacra", "OpenPitch"],
    ["Official MCP or MCP-compatible access", "✓", "✓", "✓", "◐", "✓", "◐", "✓"],
    ["No proprietary data subscription required", "◐", "✗", "✗", "✗", "✗", "✗", "✓"],
    ["Open-source server/code", "✓", "✗", "✗", "✗", "✓", "✗", "✓"],
    ["Private-company/startup data", "✗", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["Private-company revenue/ARR data", "✗", "✗", "◐", "◐", "✗", "✓", "✓"],
    ["Per-figure confidence/contradictions", "✗", "◐", "◐", "✗", "✗", "◐", "✓"],
    ["Best strategic read", "Open/MCP distribution threat or integration partner", "Direct data-platform threat", "Most direct agent-native incumbent", "Diligence-grade incumbent", "Wrapper, not full product", "Revenue-data threat", "Free/open transparent wedge"],
]


# ── Sheet 6: OSS + free direct landscape ─────────────────────────────────────

OSS_FREE_DIRECT = [
    [
        "Project/product",
        "Type",
        "Directness",
        "Open-source status",
        "Free/no-key/free-account status",
        "MCP/agent-native status",
        "Data sources used",
        "Funding/valuation support",
        "ARR/revenue support",
        "Provenance/confidence support",
        "Required keys/accounts",
        "Maintenance signal",
        "Social traction signal",
        "OpenPitch threat level",
        "Source URLs",
    ],
    [
        "Exa Company Researcher",
        "Open-source company research app",
        "near-direct",
        "MIT repo; ~1.5k GitHub stars at verification",
        "Requires Exa API key and Anthropic key; optional YouTube/GitHub keys",
        "Not MCP in this repo; social post references related Exa Company Researcher MCP tool",
        "Exa search over company website, LinkedIn, funding details, Crunchbase/PitchBook/Tracxn profiles, news, social, GitHub",
        "✓ funding lookups",
        "◐ public web/10-K only, not private ARR registry",
        "◐ source links through Exa/search results, no explicit confidence model",
        "EXA_API_KEY, ANTHROPIC_API_KEY",
        "Active-ish; 98 commits shown in GitHub UI",
        "X posts from Exa and others mention open-source/MCP/company researcher; social verification public snippets only",
        "High as open company-research proof",
        "https://github.com/exa-labs/company-researcher | https://x.com/ExaAILabs/status/1876740660738363644 | https://x.com/ExaAILabs/status/1915154310565921075",
    ],
    [
        "Octagon Funding Data MCP",
        "MCP funding/private-market intelligence service",
        "direct",
        "MCP server repo MIT; service/API is proprietary",
        "MCP.so says free account + API key; deeper free-tier limits not verified",
        "✓ MCP; Octagon also has broader market-data MCP repo",
        "Octagon private/market intelligence; funding rounds, valuations, investor activity, market trends",
        "✓",
        "✗/◐ not positioned as ARR product",
        "◐ source-backed language in main repo; no OpenPitch-style confidence model visible",
        "Octagon account/API key",
        "GitHub repo visible; MCP directory listings",
        "Public MCP directories and articles; X-specific verification not found publicly",
        "High for funding/valuation MCP",
        "https://github.com/OctagonAI/octagon-mcp-server | https://mcp.so/server/octagon-funding-data-mcp/OctagonAI",
    ],
    [
        "Sieve MCP",
        "Startup due-diligence MCP",
        "direct",
        "GitHub repo visible for lmwharton/sieve-mcp",
        "Glama connector; free-account/limits not verified from public snippets",
        "✓ MCP connector",
        "Real-time web research, evidence verification, IMPACT-X due diligence dimensions",
        "◐ diligence/funding context, not full funding DB",
        "✗ no ARR database visible",
        "✓ evidence-typed findings: Documented / Discovered / Inferred / Missing",
        "Connector/account requirements not fully verified publicly",
        "Public repo and Reddit/Glama listing",
        "Reddit r/mcp launch/connector post visible",
        "High for VC due-diligence workflow",
        "https://github.com/lmwharton/sieve-mcp | https://glama.ai/mcp/connectors?query=startup | https://www.reddit.com/r/mcp/comments/1sfd21f/sieve_aipowered_startup_due_diligence_screen_any/",
    ],
    [
        "Intelica MCP",
        "Competitive intelligence API/MCP for agents",
        "near-direct",
        "Docs public; code openness unclear from current verification",
        "Per-call x402 pricing claimed; public demo/API docs visible",
        "✓ MCP and A2A positioned publicly",
        "Competitive intelligence, market context, venture-screening modes",
        "◐ venture screening context",
        "✗ no private ARR database visible",
        "◐ structured JSON; source/provenance depth not verified",
        "Payment token / x402 for paid calls",
        "Recent dev.to launch/positioning posts",
        "Recent social/blog-style posts; no official X verification completed",
        "Medium as agent-native CI substitute",
        "https://dev.to/trustboost/the-best-competitive-intelligence-api-for-autonomous-ai-agents-2026-23md | https://github.com/teodorofodocrispin-cmyk/Intelica-docs",
    ],
    [
        "CompanyLens MCP",
        "Official-source corporate intelligence MCP",
        "near-direct",
        "MIT package/repo shown in Glama listing",
        "npx install; backend API default public URL; free limits not verified",
        "✓ MCP for Claude/Claude Code/Cursor/Windsurf",
        "SEC EDGAR, UK Companies House, OpenSanctions, USAspending.gov, SAM.gov, CourtListener",
        "✗ private funding not core",
        "✓ public-company revenue/XBRL only",
        "✓ official-source attribution; sanctions match confidence",
        "Optional custom API URL; no paid key found in public listing",
        "Package/repo links via Glama; maintenance C in directory",
        "No useful X signal found; registry discovery",
        "Medium as official-source MCP/source partner",
        "https://glama.ai/mcp/servers/diplv/companylens-mcp",
    ],
    [
        "FounderSignal MCP",
        "Founder/startup demand-signal MCP",
        "near-direct",
        "MCP listing public; repo/license not fully verified",
        "Free/account model not verified",
        "✓ MCP",
        "Crunchbase, Google Trends, LinkedIn, Meta Ads Library, TikTok, software reviews",
        "◐ Crunchbase/funding signals",
        "✗ no ARR database visible",
        "◐ source mix visible; confidence model not verified",
        "Likely external API/source requirements",
        "Appears recent in Glama",
        "No official X signal verified; registry discovery",
        "Medium for signal discovery",
        "https://glama.ai/mcp/servers/carlosalvite/foundersignal-mcp",
    ],
    [
        "Crunchbase MCP Server (Cyreslab)",
        "Open-source MCP wrapper over Crunchbase",
        "direct",
        "Open-source GitHub repo; small star/fork count",
        "Requires Crunchbase API/key/access",
        "✓ MCP",
        "Crunchbase company details, funding, acquisitions, people",
        "✓",
        "✗",
        "✗ thin wrapper; no OpenPitch-style provenance/confidence",
        "Crunchbase API key/access",
        "Small repo, low activity",
        "MCP directories list it; no meaningful social traction found",
        "Medium as easy wrapper pattern",
        "https://github.com/Cyreslab-AI/crunchbase-mcp-server | https://mcp.so/server/crunchbase/Cyreslab-AI",
    ],
    [
        "ThomasJanssen Crunchbase MCP / Bright Data variant",
        "Small Crunchbase/company-data MCP",
        "near-direct",
        "GitHub repo/listing public; tiny repo",
        "Uses Bright Data for retrieval per Glama listing; account/key likely required",
        "✓ MCP",
        "Crunchbase/company data via Bright Data",
        "◐",
        "✗",
        "✗ no confidence/provenance layer visible",
        "Bright Data / retrieval setup likely",
        "Tiny repo/listing",
        "No meaningful social traction found",
        "Low-medium wrapper threat",
        "https://github.com/ThomasJanssen-tech/MCP_Server | https://glama.ai/mcp/servers/ThomasJanssen-tech/MCP_Server",
    ],
    [
        "Tavily / Agentic Company Researcher",
        "Open-source multi-agent company research app",
        "near-direct",
        "GitHub repo visible",
        "Likely search/API dependent; exact free limits not verified",
        "Agentic app, not necessarily MCP",
        "Web/news/company info via Tavily-style search pipeline",
        "◐ public research only",
        "✗",
        "◐ report citations possible; no structured confidence model verified",
        "Search/LLM keys likely",
        "Public demo/repo visible",
        "No X verification completed",
        "Medium for company research UX",
        "https://github.com/guy-hartstein/company-research-agent",
    ],
    [
        "Apify company-research actors",
        "Paid/free actor marketplace for company research/scraping",
        "near-direct",
        "Actors generally not open data; code openness varies",
        "Apify credits/pricing; may have free trial/usage",
        "Agent/actor APIs; MCP ecosystem possible through Apify",
        "LinkedIn, PitchBook, Crunchbase, news, public web depending on actor",
        "✓/◐",
        "◐ public/reported only",
        "◐ inline summaries; no standardized confidence",
        "Apify account/credits",
        "Marketplace actors with usage stats",
        "No X verification completed",
        "Medium as commodity scraper substitute",
        "https://apify.com/pratikdani/company-research-analysis-agent",
    ],
    [
        "ROSS Index / Runa",
        "Open dataset/index for trending OSS startups",
        "adjacent",
        "GitHub dataset public",
        "Free public dataset/page",
        "✗ no MCP",
        "GitHub star growth for OSS startup repos",
        "✗",
        "✗",
        "✓ transparent scoring, not financial provenance",
        "None",
        "Maintained index/dataset",
        "No X verification completed",
        "Low-medium for OSS startup discovery",
        "https://github.com/RunaCapital/ROSS-Index | https://runacap.com/ross-index/",
    ],
    [
        "LF AI Landscape",
        "Open AI/data ecosystem landscape",
        "adjacent",
        "GitHub repo public",
        "Free public data",
        "✗ no MCP",
        "Open-source AI/data projects with GitHub/funding/market metadata",
        "◐ funding/market metadata",
        "✗",
        "◐ data-source transparency but no confidence model",
        "None",
        "Public repo",
        "No X verification completed",
        "Low-medium source/partner",
        "https://github.com/lfai/lfai-landscape",
    ],
    [
        "OpenVC / findfunding.vc",
        "Investor-side free/open VC databases",
        "adjacent",
        "OpenVC partially open/community; findfunding.vc open data posture",
        "Free access",
        "✗ no MCP",
        "Investor/fund profiles, check sizes, thesis/contact metadata",
        "✗ investor side, not company rounds",
        "✗",
        "✗",
        "None for basic use",
        "Active/community resources",
        "No X verification completed",
        "Low; complement/source partner",
        "https://www.openvc.app/ | https://www.findfunding.vc/",
    ],
    [
        "YC OSS company lists / yc-oss",
        "Maintained startup/company directories",
        "source/partner",
        "Public GitHub/list pages",
        "Free public data",
        "✗ no MCP",
        "YC/open-source startup lists and metadata",
        "✗",
        "✗",
        "✗ list data only",
        "None",
        "Maintained lists vary",
        "No X verification completed",
        "Low source input",
        "https://github.com/yc-oss/open-source-companies | https://www.ycombinator.com/companies/industry/open-source",
    ],
]


# ── Sheet 7: research backlog ────────────────────────────────────────────────

RESEARCH_BACKLOG = [
    ["Candidate", "Why it may matter", "Current status", "Next verification step", "Source URL"],
    ["StartuPage", "Founder-first startup directory claims verified revenue/fundraising/matching; possible free Crunchbase alternative", "Not fully verified; source appears SEO/listicle-like", "Inspect official product and data model", "https://startupa.ge/alternatives/crunchbase"],
    ["TopStartups / Growth List", "AI-startup lists with funding/lead data; SEO/content threat", "Known adjacent, not deeply verified", "Verify pricing, data freshness, source transparency", "https://growthlist.co/ai-startups/"],
    ["Apify Crunchbase bulk scrapers", "Commodity extraction of Crunchbase-like data", "Known actor category, not a product-level competitor", "Verify actor pricing and fields if needed", "https://apify.com/"],
    ["LinkedIn Intelligence MCP", "Company/person research through LinkedIn/RapidAPI sources", "Adjacent; access/ToS sensitive", "Keep as source-layer watch item only", "https://glama.ai/mcp/servers/southleft/linkedin-mcp"],
    ["Google Research / Exa / Tavily / Perplexity MCP", "Generic research MCPs can answer company questions live", "Near-direct only when paired with prompts; no committed data", "Document as research substrate, not competitor", "https://glama.ai/mcp/servers/search/a-guide-to-conducting-thorough-web-based-research"],
    ["Regional MENA registries / free-zone registries", "Possible source layer for MENA segment", "Not competitor products", "Evaluate as source adapters later", "ADGM / DIFC / DMCC public registries"],
    ["X/social verification for non-Exa tools", "Launch/traction signals could reveal active adoption", "Mostly blocked/noisy from public search; Exa had usable X snippets", "Use official accounts or product pages if accessible", "https://x.com/"],
]


# ── Sheet 8: adjacent watchlist ──────────────────────────────────────────────

ADJACENT = [
    ["Company/tool", "Why it matters", "Threat level", "OpenPitch implication"],
    ["Preqin", "Alternative assets/private-market data, strong institutional footprint; Bigdata.com partnership brings agentic AI access.", "Medium", "Not startup-ARR focused, but reinforces AI/private-market data trend."],
    ["Carta Data Desk", "Unique private-company/fund dataset from cap table/fund admin footprint; publishes aggregated private-market insights.", "Medium", "Data access is proprietary; OpenPitch should not imply equivalent source depth."],
    ["Beauhurst", "Strong private-company data for UK/Germany, thousands of sources, quote pricing.", "Low-medium", "Regional private-company platforms can own geography; OpenPitch wins on AI focus/no-key."],
    ["PrivCo", "Private-company financial data/research incumbent.", "Medium", "Relevant if OpenPitch leans into private-company revenue intelligence."],
    ["AlphaSense / Tegus", "Expert calls, filings, broker/research workflow; increasingly AI-driven.", "Low-medium", "Different workflow, but competes for research budget and trust."],
    ["EquityZen / Forge", "Secondary-market private-company access and research partnerships.", "Low", "Not a data-layer competitor, but source of valuation/research context."],
    ["OpenVC / findfunding.vc", "Investor-side free/open databases.", "Low", "Complementary source/partner; not company metrics."],
    ["TopStartups / Growth List", "Startup discovery/list products with AI-startup filters.", "Low", "SEO/content competitors; lack provenance/confidence/MCP."],
    ["Apify fundraising scrapers", "Commodity scraping plus MCP/payment rails.", "Low-medium", "Funding scraping is not a moat; reconciliation/provenance is."],
]


# ── Sheet 7: positioning ─────────────────────────────────────────────────────

POSITIONING = [
    ["Where OpenPitch COMPETES", "Where OpenPitch should NOT claim to compete"],
    ["Zero-cost, no-key, local MCP install", "Institutional verified deal terms"],
    ["AI-startup-specific ARR/valuation/funding first look", "Millions-company coverage breadth"],
    ["Per-figure provenance, confidence, and history", "Decades of private-market history"],
    ["Public-source discrepancy / contradiction surfacing", "Human analyst services and market maps"],
    ["Open data/files and auditable corrections", "People/founder graph depth like Harmonic"],
    ["Composable event layer for agents", "Official paid-data API breadth from incumbents"],
    ["MENA AI segment with honest lower-confidence caveat", "MENA verified funding depth like MAGNiTT"],
    ["", ""],
    ["Revised positioning statement", "The free, open, no-key intelligence layer for AI-startup metrics that agents can call — every number sourced, confidence-scored, and corrected in public."],
    ["Launch thesis", "A dense 10-company seed with real contradictions will beat a thin 50-company spreadsheet. Ship proof, then scale coverage."],
]


# ── Sheet 8: source notes ────────────────────────────────────────────────────

SOURCES = [
    ["Provider", "Source type", "Source used", "Verified claims used in workbook", "URL"],
    ["PitchBook", "official docs", "Official product/data pages + public AI capability references", "Private-market data platform; AI capabilities; API/data feed; pricing not public.", "https://pitchbook.com/"],
    ["CB Insights", "official docs", "Official Team of Agents and Microsoft 365 Copilot connector pages", "Agent workforce / Copilot connector; enterprise posture.", "https://www.cbinsights.com/team-of-agents/"],
    ["Crunchbase", "official docs", "Official Predictions and Insights post/support pages", "AI-powered predictions/insights built on private-company data; API access.", "https://about.crunchbase.com/blog/crunchbase-predictions-and-insights"],
    ["Dealroom", "official docs", "Official home/about/API/product-release pages", "Startup ecosystem data; AI harvesting; API; MCP support in release notes.", "https://dealroom.co/"],
    ["Tracxn", "official docs", "Official home/pricing pages", "Private-market intelligence; Lite/free limited; API/datadump add-ons.", "https://tracxn.com/pricing"],
    ["Harmonic", "official docs", "Official home page", "35M+ companies / 195M+ profiles; startup sourcing graph and alerts.", "https://harmonic.ai/"],
    ["MAGNiTT", "official docs", "Official home/pricing/research pages", "Emerging-market private-capital data across MENA/Africa/Turkey/Pakistan/SEA.", "https://magnitt.com/"],
    ["Wamda", "official docs", "Official home/research pages", "MENA ecosystem media/research, investment reports.", "https://www.wamda.com/research"],
    ["Sacra", "official docs + third-party listing", "Official home/API pages + third-party pricing listing for plan detail", "Private-company research; revenue/company API; pricing/MCP detail not fully first-party public.", "https://sacra.com/"],
    ["OpenBB", "official docs", "Official MCP docs", "OpenBB MCP server/workspace tools for financial workflows.", "https://docs.openbb.co/odp/python/extensions/interface/openbb-mcp"],
    ["AI Funding Tracker", "official site", "Official website", "AI-startup funding tracker, weekly/monthly funding roundup style.", "https://aifundingtracker.com/"],
    ["Crunchbase MCP", "official repo + MCP registry", "GitHub repository and MCP.so listing", "Open-source MCP server exposing Crunchbase company/funding data; requires Crunchbase access.", "https://github.com/Cyreslab-AI/crunchbase-mcp-server"],
    ["Exa Company Researcher", "official repo + social media", "GitHub repo and public X snippets", "Open-source company research app; funding/Crunchbase/PitchBook/Tracxn lookups; Exa/Anthropic keys required.", "https://github.com/exa-labs/company-researcher"],
    ["Octagon Funding MCP", "official repo + MCP registry", "GitHub repo and MCP.so listing", "Funding rounds, valuations, investor activity, market trends; free account/API key per MCP.so.", "https://mcp.so/server/octagon-funding-data-mcp/OctagonAI"],
    ["Sieve MCP", "official repo + MCP registry + social media", "GitHub, Glama, Reddit r/mcp post", "Startup due diligence MCP; evidence-typed findings and IMPACT-X score.", "https://github.com/lmwharton/sieve-mcp"],
    ["Intelica MCP", "public docs/social article", "Dev.to post and docs repo", "Competitive intelligence API/MCP/A2A positioning; per-call pricing; not open-data equivalent.", "https://dev.to/trustboost/the-best-competitive-intelligence-api-for-autonomous-ai-agents-2026-23md"],
    ["CompanyLens MCP", "MCP registry + package/repo links", "Glama listing", "Official-source corporate intelligence MCP using EDGAR, Companies House, OpenSanctions, USAspending, SAM.gov, CourtListener.", "https://glama.ai/mcp/servers/diplv/companylens-mcp"],
    ["FounderSignal MCP", "MCP registry", "Glama listing", "Founder/startup signal MCP using Crunchbase, trends, ads, LinkedIn, software reviews.", "https://glama.ai/mcp/servers/carlosalvite/foundersignal-mcp"],
    ["ThomasJanssen Crunchbase MCP", "MCP registry + repo", "Glama listing and GitHub repo", "Small MCP server for Crunchbase/company data through Bright Data-style retrieval.", "https://glama.ai/mcp/servers/ThomasJanssen-tech/MCP_Server"],
    ["Tavily / Agentic Company Researcher", "official repo", "GitHub repo", "Multi-agent company research reports; search/LLM dependent, no committed metric database.", "https://github.com/guy-hartstein/company-research-agent"],
    ["Apify company research actors", "product marketplace", "Apify actor pages", "Company research and Crunchbase/funding scraping agents; paid/free credits, not open data.", "https://apify.com/pratikdani/company-research-analysis-agent"],
    ["ROSS Index", "dataset repo + official page", "GitHub dataset and Runa page", "Open-source startup index by GitHub star growth; discovery signal, not financial metrics.", "https://github.com/RunaCapital/ROSS-Index"],
    ["LF AI Landscape", "dataset repo", "GitHub repo", "Open-source AI/data ecosystem landscape with GitHub/funding/market metadata.", "https://github.com/lfai/lfai-landscape"],
    ["OpenVC / findfunding.vc", "product/data sites", "Public sites", "Investor-side free/open databases; complement, not company metrics competitor.", "https://www.openvc.app/"],
    ["YC OSS / yc-oss", "dataset repo + product page", "GitHub list and YC company directory", "OSS startup/company lists; source input rather than product competitor.", "https://github.com/yc-oss/open-source-companies"],
    ["Free-account verification", "product account", "Account-gated verification policy", "No paid/card access used. Products requiring signup/API keys are classified from public docs/listings unless a free account is feasible in a later verification pass.", "n/a"],
    ["Preqin", "official docs", "Official site + Bigdata.com partnership page", "Alternative-assets/private-markets data; agentic AI access via partner for licensed users.", "https://www.preqin.com/"],
    ["Carta", "official docs", "Official Data Desk/About Our Data pages", "Aggregated/anonymized insights from Carta private-market customer data.", "https://carta.com/data/about/"],
    ["Beauhurst", "official docs", "Official home/pricing pages", "UK/Germany private-company data platform; quote pricing.", "https://www.beauhurst.com/"],
]


def _format_summary(ws) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = NOTE_FILL
    ws["A1"].alignment = LEFT
    _style_header(ws, 4, row=2, fill=SECTION_FILL)
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=1).font = Font(bold=True, color=NAVY)
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = LEFT
    ws.freeze_panes = "A3"


def _format_positioning(ws) -> None:
    for row in range(1, ws.max_row + 1):
        if row in (1, 10, 11):
            for col in range(1, 3):
                ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF" if row == 1 else NAVY)
                ws.cell(row=row, column=col).fill = HDR_FILL if row == 1 else NOTE_FILL


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = _write_sheet(wb, "Executive Summary", EXECUTIVE_SUMMARY, [24, 70, 18, 28], freeze="A3")
    _format_summary(summary)

    _write_sheet(wb, "Core Overview", CORE_OVERVIEW, [28] + [18] * len(CORE_COMPETITORS))
    _write_sheet(wb, "Feature Matrix", FEATURE_MATRIX, [30] + [13] * len(CORE_COMPETITORS))
    _write_sheet(wb, "Pricing & Access", PRICING, [22, 32, 14, 30, 48], freeze="A2")
    _write_sheet(wb, "Agent Open Landscape", AGENT_OPEN, [30, 22, 22, 22, 22, 22, 22, 28])
    _write_sheet(
        wb,
        "OSS + Free Direct Landscape",
        OSS_FREE_DIRECT,
        [24, 24, 14, 30, 30, 28, 42, 18, 18, 30, 26, 24, 30, 20, 58],
        freeze="D2",
    )
    _write_sheet(wb, "Research Backlog", RESEARCH_BACKLOG, [30, 50, 32, 44, 56], freeze="A2")
    _write_sheet(wb, "Adjacent Watchlist", ADJACENT, [24, 58, 18, 54], freeze="A2")
    positioning = _write_sheet(wb, "Positioning", POSITIONING, [58, 66], freeze="A2")
    _format_positioning(positioning)
    _write_sheet(wb, "Source Notes", SOURCES, [24, 22, 42, 64, 58], freeze="A2")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = LEFT if len(str(cell.value or "")) > 3 else CENTER
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 36 if row_idx > 1 else 28

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
