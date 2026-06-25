from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/Users/mohamedibnomer/openpitch/docs/OpenPitch-PM-Workbook.xlsx"

NAVY = "1F2A44"; ACCENT = "2E7D6F"; LIGHT = "EAF2EF"; ZEBRA = "F6F8FA"
HDR = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE = Font(name="Arial", bold=True, color=NAVY, size=16)
SUB = Font(name="Arial", italic=True, color="666666", size=10)
BOLD = Font(name="Arial", bold=True, color=NAVY, size=11)
BODY = Font(name="Arial", size=10)
BLUE = Font(name="Arial", size=10, color="0000FF")   # inputs
BLACK = Font(name="Arial", size=10, color="000000")  # formulas
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=ACCENT)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CT = Alignment(horizontal="center", vertical="center")
LT = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()


def hrow(ws, r, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h); cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CT; cell.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w


def grid(ws, r0, r1, c1):
    for r in range(r0, r1 + 1):
        for c in range(1, c1 + 1):
            cell = ws.cell(r, c); cell.border = BORDER
            if not cell.font or cell.font.name != "Arial":
                cell.font = BODY
            cell.alignment = LT
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)


# ── Cover ────────────────────────────────────────────────────────────────────
ws = wb.active; ws.title = "Cover"
ws.column_dimensions["A"].width = 4; ws.column_dimensions["B"].width = 40; ws.column_dimensions["C"].width = 60
ws["B2"] = "OpenPitch — PM Workbook"; ws["B2"].font = TITLE
ws["B3"] = "KPIs · Launch Gates · Cost Model · Risk Register · Roadmap"; ws["B3"].font = SUB
ws["B4"] = "Owner: Mohamed Abdulhadi   ·   Updated: 2026-06-14"; ws["B4"].font = SUB
rows = [
    ("Tab", "Purpose"),
    ("KPI Tracker", "North-star + activation/trust/usage/community metrics with targets and weekly star milestones"),
    ("Launch Gates", "Go/no-go checklist (auto % complete) — must all be Done before public launch"),
    ("Cost Model", "Proves the zero-cost claim: free-tier capacity vs daily demand, and cost-if-paid"),
    ("Risk Register", "Likelihood × impact scoring with mitigations and owners"),
    ("Roadmap", "Time-boxed phases and milestones"),
]
r = 6
for a, b in rows:
    ws.cell(r, 2, a).font = BOLD if r == 6 else BODY
    ws.cell(r, 3, b).font = BOLD if r == 6 else BODY
    ws.cell(r, 2).alignment = LT; ws.cell(r, 3).alignment = LT
    r += 1
ws.cell(r + 1, 2, "Legend").font = BOLD
ws.cell(r + 2, 2, "Blue text = input you change").font = BLUE
ws.cell(r + 3, 2, "Black text = formula (auto-calculated)").font = BLACK

# ── KPI Tracker ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("KPI Tracker")
hrow(ws, 1, ["Category", "Metric", "Definition", "Target", "Current", "Gap"], [16, 30, 44, 14, 12, 12])
kpis = [
    ("North Star", "Weekly returning MCP users", "Distinct users who query OpenPitch in-agent ≥1×/wk", 50, 0),
    ("Activation", "Time-to-first-answer (sec)", "README → first sourced answer", 60, 0),
    ("Activation", "Clean-machine install success %", "Fresh-machine MCP install works", 1.0, 0),
    ("Trust", "% metrics with full provenance", "value+source+confidence+as_of", 1.0, 1.0),
    ("Trust", "Correction issues resolved", "Accepted public corrections", 10, 0),
    ("Trust", "Confirmed:contradicted ratio", "Claims later confirmed vs contradicted", 3, 0),
    ("Usage", "MCP tool calls / week", "Across demos + real users", 500, 0),
    ("Usage", "Event-feed / recipe consumers", "Downstream integrations", 10, 0),
    ("Usage", "Dashboard visits / week", "Static dashboard traffic", 1000, 0),
    ("Community", "GitHub stars (30-day)", "Launch-window stars", 5000, 0),
    ("Community", "Contributors", "Distinct PR authors", 15, 0),
    ("Community", "PRs (data/sources/adapters)", "Community contributions", 25, 0),
]
r = 2
for cat, m, d, t, cur in kpis:
    ws.cell(r, 1, cat); ws.cell(r, 2, m); ws.cell(r, 3, d)
    tc = ws.cell(r, 4, t); vc = ws.cell(r, 5, cur)
    tc.font = BLUE; vc.font = BLUE
    gc = ws.cell(r, 6, f"=D{r}-E{r}"); gc.font = BLACK
    pct = (t <= 1.0 and t > 0) or "%" in m
    fmt = "0.0%" if (isinstance(t, float) and t <= 1.0) else "#,##0"
    for col in (4, 5, 6):
        ws.cell(r, col).number_format = fmt
    r += 1
grid(ws, 1, r - 1, 6)
# weekly star milestones
ms = r + 2
ws.cell(ms, 1, "Weekly star milestones").font = BOLD
hrow(ws, ms + 1, ["Week", "Target", "Actual", "On track?"], None)
weeks = [("Week 1", 1000), ("Week 2", 2500), ("Week 3", 4000), ("Week 4", 5000)]
rr = ms + 2
for wlabel, target in weeks:
    ws.cell(rr, 1, wlabel); tc = ws.cell(rr, 2, target); tc.font = BLUE; ac = ws.cell(rr, 3, 0); ac.font = BLUE
    ws.cell(rr, 4, f'=IF(C{rr}>=B{rr},"Yes","No")').font = BLACK
    rr += 1
grid(ws, ms + 1, rr - 1, 4)

# ── Launch Gates ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Launch Gates")
ws["A1"] = "Launch readiness — all gates must be Done before public launch"; ws["A1"].font = BOLD
ws["A2"] = "% complete:"; ws["A2"].font = BOLD
pcell = ws["B2"]; pcell.font = BLACK
hrow(ws, 4, ["#", "Gate", "Required", "Status", "Owner", "Target date", "Notes"], [4, 34, 10, 14, 12, 14, 40])
gates = [
    ("Credible seed dataset (10 cos, 3+ sourced metrics)", "Yes"),
    ("Working MCP read path (tools return + no-data handled)", "Yes"),
    ("Provenance on every metric (source+confidence+as_of)", "Yes"),
    ("2–3 strong public-source contradictions", "Yes"),
    ("Correction workflow + issue template", "Yes"),
    ("Methodology + Data Policy published", "Yes"),
    ("Static dashboard / company cards", "Yes"),
    ("README demo asset (GIF/screens)", "Yes"),
    ("Clean-machine install test", "Yes"),
    ("MCP directory submissions prepared", "Before campaign"),
]
dv = DataValidation(type="list", formula1='"Not started,In progress,Done"', allow_blank=True)
ws.add_data_validation(dv)
r = 5
for i, (g, req) in enumerate(gates, 1):
    ws.cell(r, 1, i); ws.cell(r, 2, g); ws.cell(r, 3, req)
    sc = ws.cell(r, 4, "Not started"); sc.font = BLUE; dv.add(sc)
    ws.cell(r, 5, ""); ws.cell(r, 6, ""); ws.cell(r, 7, "")
    r += 1
grid(ws, 4, r - 1, 7)
pcell.value = f'=COUNTIF(D5:D{r-1},"Done")/COUNTA(D5:D{r-1})'
pcell.number_format = "0%"

# ── Cost Model ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Cost Model")
ws["A1"] = "Zero-cost capacity model — does the free tier cover a daily 50-company refresh?"; ws["A1"].font = BOLD
hrow(ws, 3, ["Input (blue = change me)", "Value", "Unit", "Source / note"], [38, 14, 14, 46])
inputs = [
    ("Companies tracked", 50, "count", "PRD scope"),
    ("Source items per company", 15, "items", "news+podcast+edgar+web caps"),
    ("Batch size (items per LLM call)", 15, "items", "extract_claims_batch"),
    ("Gemini 2.5-flash free requests/day", 20, "req/day", "Observed live (quota error 2026-06-14)"),
    ("Gemini 2.0-flash free requests/day", 200, "req/day", "Est.; own daily quota"),
    ("Gemini flash-lite free requests/day", 200, "req/day", "Est.; own daily quota"),
    ("Groq Whisper free requests/day", 2000, "req/day", "Groq free tier (transcription)"),
    ("Paid price per 1M input tokens (Flash)", 0.30, "USD", "Est. if ever paid"),
    ("Avg tokens per batch call", 6000, "tokens", "~15 items × ~400 tok"),
]
r = 4
labelrow = {}
for name, val, unit, src in inputs:
    ws.cell(r, 1, name); vc = ws.cell(r, 2, val); vc.font = BLUE
    ws.cell(r, 3, unit); ws.cell(r, 4, src).font = SUB
    labelrow[name] = r
    if isinstance(val, float):
        vc.number_format = "$#,##0.00"
    r += 1
grid(ws, 3, r - 1, 4)

calc0 = r + 1
ws.cell(calc0, 1, "Calculations").font = BOLD
hrow(ws, calc0 + 1, ["Metric", "Value", "Unit", "Note"], None)
ci = labelrow["Companies tracked"]; si = labelrow["Source items per company"]; bi = labelrow["Batch size (items per LLM call)"]
g25 = labelrow["Gemini 2.5-flash free requests/day"]; g20 = labelrow["Gemini 2.0-flash free requests/day"]; gl = labelrow["Gemini flash-lite free requests/day"]
ppm = labelrow["Paid price per 1M input tokens (Flash)"]; tpc = labelrow["Avg tokens per batch call"]
cr = calc0 + 2
def calc(label, formula, fmt, note=""):
    global cr
    ws.cell(cr, 1, label); fc = ws.cell(cr, 2, formula); fc.font = BLACK; fc.number_format = fmt
    ws.cell(cr, 3, ""); ws.cell(cr, 4, note).font = SUB
    cr += 1
    return cr - 1
row_calls_co = calc("LLM calls per company", f"=CEILING(B{si}/B{bi},1)", "#,##0", "batched")
row_demand = calc("Total extraction calls/day (demand)", f"=B{ci}*B{row_calls_co}", "#,##0", "")
row_cap = calc("Free capacity/day (3-model rotation)", f"=B{g25}+B{g20}+B{gl}", "#,##0", "each model has its own daily quota")
row_head = calc("Headroom (capacity − demand)", f"=B{row_cap}-B{row_demand}", "#,##0", "positive = fits in free tier")
row_fits = calc("Fits in free tier?", f'=IF(B{row_head}>=0,"YES — $0","NO — add provider/day-split")', "General", "")
row_tokday = calc("Tokens/day used", f"=B{row_demand}*B{tpc}", "#,##0", "")
row_costpaid = calc("Cost/day IF paid (not needed)", f"=B{row_tokday}/1000000*B{ppm}", "$#,##0.00", "reference only")
row_costmo = calc("Cost/month IF paid (not needed)", f"=B{row_costpaid}*30", "$#,##0.00", "reference only")
grid(ws, calc0 + 1, cr - 1, 4)

# ── Risk Register ────────────────────────────────────────────────────────────
ws = wb.create_sheet("Risk Register")
hrow(ws, 1, ["#", "Risk", "Likelihood (1-5)", "Impact (1-5)", "Score", "Mitigation", "Owner", "Status"], [4, 34, 14, 12, 9, 40, 12, 14])
risks = [
    ("Data inaccuracy damages credibility", 3, 5, "Provenance + confidence + ranges; never assert false precision; corrections"),
    ("Audience ≠ pain (stars ≠ PMF)", 4, 5, "Re-anchor ICP to AI-builder grounding; talk to first 50 users weekly"),
    ("Free-tier LLM quota too low", 3, 4, "Batching + multi-provider rotation; Cerebras for tokens"),
    ("Legal/defamation on contradictions", 2, 5, "Neutral 'public-source discrepancy' language; takedown policy"),
    ("Source ToS / scraping blocks", 3, 3, "Official APIs/RSS first; respect robots; conservative posture"),
    ("Low launch adoption / HN flop", 3, 4, "Portfolio of shots; contradiction hook; second-chance pool"),
    ("Funded OSS rival pivots (e.g. OpenBB)", 2, 4, "Speed + niche depth; interoperate, not only compete"),
    ("Single-maintainer bandwidth", 4, 3, "Keep scope narrow; community contributors; automation"),
]
r = 2
for i, (rk, l, im, mit) in enumerate(risks, 1):
    ws.cell(r, 1, i); ws.cell(r, 2, rk)
    lc = ws.cell(r, 3, l); lc.font = BLUE; ic = ws.cell(r, 4, im); ic.font = BLUE
    sc = ws.cell(r, 5, f"=C{r}*D{r}"); sc.font = BLACK; sc.alignment = CT
    ws.cell(r, 6, mit); ws.cell(r, 7, ""); st = ws.cell(r, 8, "Open"); st.font = BLUE
    r += 1
grid(ws, 1, r - 1, 8)

# ── Roadmap ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Roadmap")
hrow(ws, 1, ["Phase", "Milestone", "Status", "Target", "Notes"], [16, 40, 14, 14, 34])
road = [
    ("P0 Foundation", "Specs + data model + reconciliation engine", "Done", "", "shipped"),
    ("P1 Pipeline", "Adapters + extraction + derivation + publish", "Done", "", "shipped"),
    ("P1 Pipeline", "Storage/IO + orchestration + dashboard + MCP", "Done", "", "shipped"),
    ("P2 Data", "Verified 10-company seed + contradictions", "In progress", "", "verify launch-grade"),
    ("P2 Data", "Whisper (Groq) live + EDGAR UA fix", "In progress", "", "podcast edge"),
    ("P3 Launch prep", "LICENSE + governance + issue templates", "In progress", "", "this batch"),
    ("P3 Launch prep", "Pitch deck + case study + demo GIF", "Not started", "", "showcase"),
    ("P3 Publish", "PyPI publish + set OPENPITCH_REMOTE", "Not started", "", "uvx install"),
    ("P4 GTM", "Run 30-day growth plan", "Not started", "", "after gates pass"),
    ("P5 Scale", "A2A agent · MENA adapters · rich sources", "Not started", "", "post-PMF"),
]
dv2 = DataValidation(type="list", formula1='"Not started,In progress,Done"', allow_blank=True)
ws.add_data_validation(dv2)
r = 2
for ph, mi, st, tg, nt in road:
    ws.cell(r, 1, ph); ws.cell(r, 2, mi); sc = ws.cell(r, 3, st); sc.font = BLUE; dv2.add(sc)
    ws.cell(r, 4, tg); ws.cell(r, 5, nt)
    r += 1
grid(ws, 1, r - 1, 5)

for s in wb.sheetnames:
    wb[s].sheet_view.showGridLines = False
wb.save(OUT)
print("wrote", OUT)
